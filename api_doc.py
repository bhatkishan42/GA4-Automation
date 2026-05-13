# Version 28 — Detect "(not set)" tokens in GA4 actual values
# (was: Version 27 — Add --mode CLI flag for normal vs ignore-required)
#
# What v28 adds:
#   - A small token-detector that flags rows whose ga4_actual_value contains
#     "(not set)" (or close variants — "(not_set)", "not set", "<not set>").
#     "(not set)" is what GA4 puts in dimension columns when an event fires
#     without that parameter populated, so it's almost always a real signal
#     rather than a clean value.
#   - When detected, an annotation is appended to the row's Comments column:
#     [Note: contains (not set)] — non-destructive, so a Pass row stays Pass
#     and a Failed row stays Failed but you still see the (not set) flag.
#   - End-of-run console line shows how many rows tripped the detector.
#
# Detection happens AFTER status decision and applies on QA'd rows only —
# unchecked rows still show up as Manual Check with no annotation, since
# we haven't really evaluated them yet.

import argparse
import pandas as pd
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule

# =========================
# CONFIG — only change these
# =========================
RULES_FILE  = "ignore_required.xlsx"            # the Rules / spec workbook (now also holds GA4 values)
SHEET_NAME  = "API Confluence (Final Version)"
OUTPUT_FILE = "Api_validation_output.xlsx"

# If True, appends timestamp to output filename (keeps run history)
KEEP_HISTORY = True

# Hardcoded total events & modules per platform — UPDATE WHEN FULL SPEC CHANGES
TOTAL_EVENTS_APP    = 275   # change when full app event count changes
TOTAL_EVENTS_WEB    = 178   # change when full web event count is known
TOTAL_MODULES_APP   = 29    # change when full app module count changes
TOTAL_MODULES_WEB   = 29    # change when full web module count changes

# Marker value in ga4_check_status column that means "this row has been QA'd"
QA_CHECKED_MARKER = "checked"

# =========================
# Racing THEME COLORS (Summary sheet)
# =========================
COLOR_BG             = "15151E"
COLOR_HEADER         = "FF1E00"
COLOR_SUBHEADER      = "1E1E2A"
COLOR_ROW_ODD        = "1A1A27"
COLOR_ROW_EVEN       = "22222F"
COLOR_TEXT_PRIMARY   = "FFFFFF"
COLOR_TEXT_SECONDARY = "E8E8E8"
COLOR_TEXT_MUTED     = "C0C0C0"
COLOR_BORDER         = "3A3A4A"
COLOR_BAR_FILL       = "00B050"

# Validation sheet — traffic-light colors for the Status column
COLOR_VAL_GREEN      = "63BE7B"
COLOR_VAL_YELLOW     = "FFEB84"
COLOR_VAL_RED        = "F8696B"

BG_FILL       = PatternFill("solid", fgColor=COLOR_BG)
HEADER_FILL   = PatternFill("solid", fgColor=COLOR_HEADER)
SUB_FILL      = PatternFill("solid", fgColor=COLOR_SUBHEADER)
ROW_ODD_FILL  = PatternFill("solid", fgColor=COLOR_ROW_ODD)
ROW_EVEN_FILL = PatternFill("solid", fgColor=COLOR_ROW_EVEN)

CF_GREEN_FILL  = PatternFill("solid", fgColor=COLOR_VAL_GREEN)
CF_YELLOW_FILL = PatternFill("solid", fgColor=COLOR_VAL_YELLOW)
CF_RED_FILL    = PatternFill("solid", fgColor=COLOR_VAL_RED)

THIN_BORDER  = Side(border_style="thin", color=COLOR_BORDER)
CELL_BORDER  = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)

FONT_TITLE      = Font(name="Calibri", size=18, bold=True, color=COLOR_TEXT_PRIMARY)
FONT_SUBTITLE   = Font(name="Calibri", size=10, italic=True, color=COLOR_TEXT_MUTED)
FONT_SECTION    = Font(name="Calibri", size=12, bold=True, color=COLOR_TEXT_PRIMARY)
FONT_HEADER     = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT_PRIMARY)
FONT_LABEL      = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT_SECONDARY)
FONT_VALUE      = Font(name="Calibri", size=11, color=COLOR_TEXT_PRIMARY)
FONT_VALUE_BOLD = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT_PRIMARY)


# =========================
# NORMALIZE HELPERS
# =========================
def normalize(value):
    return str(value).strip().lower()


# Tokens GA4 emits when a parameter wasn't actually populated on an event.
# All comparisons are done case-insensitively and after stripping whitespace,
# so this list only needs the canonical lower-case forms.
NOT_SET_TOKENS = {
    "(not set)",
    "(not_set)",
    "not set",
    "<not set>",
}


def _has_not_set(raw):
    """True iff the raw cell value contains a recognised (not set) token.

    Operates on the RAW cell (pre-split, pre-normalize) so we catch the
    token regardless of which side of a pipe it's on. Returns False for
    empty/NaN cells.
    """
    if raw is None or pd.isna(raw):
        return False
    s = str(raw).lower()
    return any(t in s for t in NOT_SET_TOKENS)


def split_values(raw):
    """Split a cell value into a list of normalized tokens."""
    if raw is None or pd.isna(raw):
        return []
    s = str(raw).strip()
    if not s or s.lower() in ("nan", "none"):
        return []
    s = s.replace("\r", "").replace("\t", "")
    for sep in ("\n", ","):
        s = s.replace(sep, "|")
    return [normalize(v) for v in s.split("|") if v.strip()]


# =========================
# READ EXCEL  (Rules + GA4 values now in same file)
# =========================
def read_excel(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    df.columns = df.columns.str.strip().str.lower()

    # AUTO-FIX swapped event_name/platform
    platform_labels = {"app", "web"}
    mask = df["event_name"].astype(str).str.strip().str.lower().isin(platform_labels)
    if mask.any():
        fixed_count = int(mask.sum())
        print(f"⚠ Auto-fixing {fixed_count} rows where event_name/platform are swapped")
        swapped_events   = df.loc[mask, "platform"].copy()
        swapped_platform = df.loc[mask, "event_name"].copy()
        df.loc[mask, "event_name"] = swapped_events
        df.loc[mask, "platform"]   = swapped_platform

    df = df[df["event_name"].notna() & (df["event_name"].astype(str).str.strip() != "")]

    # Back-compat: alias old column names if present
    if "expected_value" in df.columns and "rules_expected_values" not in df.columns:
        df = df.rename(columns={"expected_value": "rules_expected_values"})
    if "ga4_expected_values" in df.columns and "ga4_actual_value" not in df.columns:
        df = df.rename(columns={"ga4_expected_values": "ga4_actual_value"})

    # Sanity-check required columns
    required_cols = {"event_name", "platform", "parameter_name",
                     "rules_expected_values", "ga4_actual_value"}
    missing = required_cols - set(df.columns)
    if missing:
        raise ValueError(
            f"Rules sheet is missing required columns: {sorted(missing)}. "
            f"Found columns: {list(df.columns)}"
        )

    has_check_status_col = "ga4_check_status" in df.columns
    if not has_check_status_col:
        print("⚠ ga4_check_status column not found — QA Completed metrics will fall back "
              "to counting all events/modules in the spec.\n")

    # Capture original column order for downstream output passthrough
    input_columns = list(df.columns)

    events = {}
    duplicates = []

    # Track input-file order of modules (first-appearance, post empty→Uncategorized).
    # This is what later drives the Summary sheet's Per-Module Breakdown ordering.
    module_order = []
    seen_modules = set()

    for _, row in df.iterrows():
        raw_mod         = row.get("module", "")
        module          = "" if pd.isna(raw_mod) else str(raw_mod).strip()
        module_label    = module if module else "Uncategorized"
        raw_event       = str(row["event_name"])
        platform        = normalize(row.get("platform", "all"))
        parameter_name  = normalize(row["parameter_name"])

        bq_column       = "" if pd.isna(row.get("bq_column",  ""))  else str(row.get("bq_column",  "")).strip()
        api_column      = "" if pd.isna(row.get("api_column", "")) else str(row.get("api_column", "")).strip()

        rules_expected  = split_values(row.get("rules_expected_values"))
        ga4_actual_raw  = row.get("ga4_actual_value")
        ga4_actual      = split_values(ga4_actual_raw)

        required        = str(row.get("required", "")).strip().lower() in ["yes", "true", "1"]

        # Per-row QA-checked flag
        if has_check_status_col:
            check_raw = row.get("ga4_check_status")
            check_str = "" if pd.isna(check_raw) else str(check_raw).strip().lower()
            qa_checked = (check_str == QA_CHECKED_MARKER.lower())
        else:
            qa_checked = True

        # Capture the raw row as a dict (for output passthrough). NaN -> ""
        raw_row = {col: ("" if pd.isna(row[col]) else row[col]) for col in input_columns}

        # First-appearance module order
        if module_label not in seen_modules:
            seen_modules.add(module_label)
            module_order.append(module_label)

        event_names = [normalize(e) for e in raw_event.split("|") if e.strip()]

        for event_name in event_names:
            # Key includes module so the same event_name under two different
            # modules doesn't collapse into one bucket. (See v26 banner.)
            key = f"{module_label}||{event_name}||{platform}"

            if key not in events:
                events[key] = {
                    "module"     : module,
                    "event_name" : event_name,
                    "platform"   : platform,
                    "params"     : []
                }

            existing_params = {p["parameter_name"] for p in events[key]["params"]}
            if parameter_name in existing_params:
                duplicates.append({
                    "module"    : module_label,
                    "event"     : event_name,
                    "platform"  : platform,
                    "parameter" : parameter_name,
                })
                continue

            # For pipe-separated events, ensure the row's event_name field shows
            # the specific event_name being validated (not the original "evt1|evt2")
            row_for_this_event = dict(raw_row)
            row_for_this_event["event_name"] = event_name

            events[key]["params"].append({
                "parameter_name" : parameter_name,
                "bq_column"      : bq_column,
                "api_column"     : api_column,
                "rules_expected" : rules_expected,
                "ga4_actual"     : ga4_actual,
                "ga4_actual_raw" : ga4_actual_raw,
                "required"       : required,
                "qa_checked"     : qa_checked,
                "raw_row"        : row_for_this_event,
            })

    if duplicates:
        print(f"\n⚠ Found {len(duplicates)} duplicate parameter definitions (kept first occurrence):")
        for d in duplicates:
            print(f"    - [{d['module']}] {d['event']} [{d['platform']}] → {d['parameter']}")
        print()

    return events, input_columns, module_order


# =========================
# VALIDATE  (row-internal compare — no external dataframe)
# =========================
def validate(events, ignore_required=False):
    results = []

    for key, config in events.items():
        module     = config["module"]
        event_name = config["event_name"]
        platform   = config["platform"]

        event_qa_checked = any(p["qa_checked"] for p in config["params"])

        event_found = any(p["ga4_actual"] for p in config["params"] if p["required"])
        if not event_found:
            event_found = any(p["ga4_actual"] for p in config["params"])

        print(f"\n▶ [{module or 'Uncategorized'}] {event_name} [{platform.upper()}]"
              f" — {'FOUND' if event_found else 'NOT FOUND'}"
              f"  | QA: {'CHECKED' if event_qa_checked else 'pending'}")

        for param in config["params"]:
            param_name      = param["parameter_name"]
            expected_values = param["rules_expected"]
            ga4_actual      = param["ga4_actual"]
            required        = param["required"]
            qa_checked      = param["qa_checked"]

            # In ignore-required mode every QA'd row is treated as required
            # so the value matcher runs on it. Original `required` is left
            # alone for any other consumer that might care.
            effective_required = True if ignore_required else required

            actual_values_sorted = sorted(set(ga4_actual))

            # Only validate rows that have been QA'd. Unchecked rows pass through
            # as "Manual Check" so they stay visible but don't get validated.
            if not qa_checked:
                status = "NOT REQUIRED"
                print(f"  - {param_name} | not QA'd yet — skipped")

            elif not effective_required:
                status = "NOT REQUIRED"
                print(f"  - {param_name} | NOT REQUIRED — skipped")

            elif not ga4_actual:
                status = "EVENT NOT FOUND" if not event_found else "PARAMETER MISSING"

            else:
                actual_set = set(ga4_actual)

                if not expected_values:
                    status = "FOUND (no expected value specified)"
                elif all(v in actual_set for v in expected_values):
                    additional = sorted([v for v in actual_values_sorted if v not in expected_values])
                    status = f"PARTIAL — additional values: {', '.join(additional)}" if additional else "PASS"
                elif any(v in actual_set for v in expected_values):
                    missing = [v for v in expected_values if v not in actual_set]
                    status  = f"PARTIAL — missing: {', '.join(missing)}"
                else:
                    status = "FAIL — no expected values found in GA4"

            if effective_required and qa_checked:
                icon = (
                    "✔" if status == "PASS"
                    else "~" if any(x in status for x in ["no expected", "PARTIAL"])
                    else "✖"
                )
                print(f"  {icon} {param_name} | {status}")

            display_status, comments = status_to_columns(status)

            # (not set) detection — annotate any QA'd row whose raw GA4 cell
            # contains a recognised "(not set)" token. We do this regardless
            # of pass/fail so that even a passing row carries the flag.
            has_not_set = qa_checked and _has_not_set(param.get("ga4_actual_raw"))
            if has_not_set:
                note = "[Note: contains (not set)]"
                comments = f"{comments} {note}".strip() if comments else note

            # Start with the full original row, then attach computed fields
            result_row = dict(param["raw_row"])
            result_row["Status"]   = display_status
            result_row["Comments"] = comments
            # Internal-only fields (dropped before writing the sheet)
            result_row["_status_internal"] = status
            result_row["_qa_checked"]      = qa_checked
            result_row["_has_not_set"]     = has_not_set
            # For the summary builders that group by these
            result_row["_module"]     = module
            result_row["_event_name"] = event_name
            result_row["_platform"]   = platform.upper()
            # _required reflects the EFFECTIVE required-ness, which is what
            # the Summary's Modules Completed tally cares about.
            result_row["_required"]   = "YES" if effective_required else "NO"

            results.append(result_row)

    return pd.DataFrame(results)


# =========================
# STATUS / COMMENTS TRANSLATOR
# =========================
def status_to_columns(internal_status):
    """
    Translate the internal status string into (display_status, comments) for
    the Validation output sheet.

    Display values (only 5):
      - Pass
      - Failed
      - Partial - Missing
      - Partial - Additional
      - Manual Check
    """
    s = internal_status

    if s == "PASS":
        return "Pass", ""

    if s == "NOT REQUIRED":
        return "Manual Check", ""

    if s == "EVENT NOT FOUND":
        return "Failed", "event not in GA4"
    if s == "PARAMETER MISSING":
        return "Failed", "no values in GA4"
    if s == "FAIL — no expected values found in GA4":
        return "Failed", "no expected values found in GA4"
    if s == "FOUND (no expected value specified)":
        return "Failed", "data exists but no expected value specified in Rules"

    if s.startswith("PARTIAL — missing: "):
        return "Partial - Missing", s.replace("PARTIAL — ", "")
    if s.startswith("PARTIAL — additional values: "):
        return "Partial - Additional", s.replace("PARTIAL — ", "")

    return "Failed", s


# =========================
# SUMMARY HELPERS
# =========================
def categorize(status):
    """Bucket the internal status for summary calculations."""
    if status == "PASS" or status == "FOUND (no expected value specified)":
        return "PASS"
    if status.startswith("PARTIAL"):
        return "PARTIAL"
    if status == "NOT REQUIRED":
        return "EXCLUDED"
    return "FAIL"


def build_metrics_for_platform(result_df, platform_label, total_events_spec,
                               total_modules_spec, module_order):
    """Compute completion-rate metrics + per-module param table for one platform.

    QA Completed metrics + Per-module table are driven by ga4_check_status:
      - QA Completed (events)  = unique events with ≥1 row marked "checked"
      - QA Completed (modules) = unique modules with ≥1 row marked "checked"
      - Per-module rows         = only show modules that have ≥1 checked row,
                                  ordered by first appearance in the input file.
    """
    df = result_df[result_df["_platform"].str.upper() == platform_label].copy()
    df["module_label"] = df["_module"].replace("", "Uncategorized").fillna("Uncategorized")
    df["bucket"]       = df["_status_internal"].apply(categorize)

    checked_df = df[df["_qa_checked"] == True]

    qa_events_completed  = checked_df["_event_name"].nunique()  if len(checked_df) else 0
    qa_modules_completed = checked_df["module_label"].nunique() if len(checked_df) else 0

    validated   = df[df["bucket"] != "EXCLUDED"]
    events_acc  = df["_event_name"].nunique() if len(df) else 0
    modules_acc = df["module_label"].nunique() if len(df) else 0

    required = df[df["_required"] == "YES"]
    modules_completed = 0
    if not required.empty:
        for _, grp in required.groupby("module_label"):
            if (grp["bucket"] == "PASS").all():
                modules_completed += 1

    metrics = {
        "Total Events"             : total_events_spec,
        "Events Accounted"         : qa_events_completed,
        "Event Completion %"       : round(qa_events_completed  / total_events_spec  * 100, 2) if total_events_spec  else 0.0,
        "Modules"                  : total_modules_spec,
        "Modules Worked On"        : qa_modules_completed,
        "Module Progress %"        : round(qa_modules_completed / total_modules_spec * 100, 2) if total_modules_spec else 0.0,
        "Modules Completed"        : modules_completed,
        "Module Completion %"      : round(modules_completed / total_modules_spec * 100, 2) if total_modules_spec else 0.0,
        "Total Modules"            : modules_acc,
        "Total Unique Events"      : events_acc,
        "Total Parameters Checked" : len(validated),
    }

    # Build per-module rows. Use sort=False on groupby so we don't fight the
    # explicit reorder we apply right after.
    rows = []
    for mod, grp in checked_df.groupby("module_label", sort=False):
        mod_validated = grp[grp["bucket"] != "EXCLUDED"]
        n  = len(mod_validated)
        p  = int((mod_validated["bucket"] == "PASS").sum())
        pp = int((mod_validated["bucket"] == "PARTIAL").sum())
        fl = int((mod_validated["bucket"] == "FAIL").sum())

        events_in_mod = grp["_event_name"].nunique()

        if n == 0:
            continue

        assert p + pp + fl == n, f"Math mismatch for {mod}: {p}+{pp}+{fl}!={n}"
        rows.append({
            "Module"     : mod,
            "Events"     : events_in_mod,
            "Parameters" : n,
            "PASS"       : p,
            "PARTIAL"    : pp,
            "FAIL"       : fl,
            "Pass %"     : round(p / n * 100, 1) if n else 0.0,
        })

    if rows:
        order_idx = {m: i for i, m in enumerate(module_order)}
        # Anything not in module_order (shouldn't happen, but be safe) sinks to
        # the end while preserving relative order.
        sentinel = len(module_order)
        rows.sort(key=lambda r: order_idx.get(r["Module"], sentinel))
        module_param_df = pd.DataFrame(rows).reset_index(drop=True)
    else:
        module_param_df = pd.DataFrame()

    return metrics, module_param_df


# =========================
# WRITE SUMMARY SHEET
# =========================
def fill_background(ws, max_row, max_col):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = BG_FILL


def section_header(ws, row, col_start, col_end, text):
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = FONT_SECTION
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if col_end > col_start:
        ws.merge_cells(start_row=row, end_row=row, start_column=col_start, end_column=col_end)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = HEADER_FILL
    ws.row_dimensions[row].height = 24


def style_cell(cell, value=None, fill=None, font=None, border=CELL_BORDER, align="center", num_format=None):
    if value is not None:
        cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if border is not None:
        cell.border = border
    if align == "center":
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif align == "left":
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if num_format:
        cell.number_format = num_format


def write_summary_sheet(wb, app_metrics, app_mod_params, web_metrics, web_mod_params,
                        module_order, mode_label="Normal"):
    sheet_name = "Summary"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    fill_background(ws, max_row=80, max_col=14)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COLOR_HEADER

    style_cell(ws.cell(row=1, column=1, value="GA4 EVENT VALIDATION  |  SUMMARY"),
               fill=HEADER_FILL, font=FONT_TITLE, border=None)
    ws.merge_cells("A1:N1")
    ws.row_dimensions[1].height = 40

    style_cell(ws.cell(row=2, column=1,
               value=(f"Generated: {datetime.now(ZoneInfo('Asia/Kolkata')).strftime('%B %d, %Y at %I:%M %p IST')}"
                      f"  |  Mode: {mode_label}")),
               fill=BG_FILL, font=FONT_SUBTITLE, border=None)
    ws.merge_cells("A2:N2")
    ws.row_dimensions[2].height = 20

    # Table 1 - QA Overview
    row = 4
    section_header(ws, row, 1, 3, "  QA OVERVIEW")
    row += 1

    style_cell(ws.cell(row=row, column=1, value=""),    fill=SUB_FILL, font=FONT_HEADER, align="left")
    style_cell(ws.cell(row=row, column=2, value="App"), fill=SUB_FILL, font=FONT_HEADER)
    style_cell(ws.cell(row=row, column=3, value="Web"), fill=SUB_FILL, font=FONT_HEADER)
    row += 1

    pct_rows_to_bar = []
    cr_metrics = [
        ("Total Events",              "Total Events",        None),
        ("QA Completed",              "Events Accounted",    None),
        ("QA Completion % (Events)",  "Event Completion %",  '0.00"%"'),
        ("__BREAK__",                 None,                  None),
        ("Modules",                   "Modules",             None),
        ("QA Completed",              "Modules Worked On",   None),
        ("Completion % (Modules)",    "Module Progress %",   '0.00"%"'),
    ]

    for idx, (label, key, fmt) in enumerate(cr_metrics):
        rfill = ROW_ODD_FILL if idx % 2 == 0 else ROW_EVEN_FILL

        if label == "__BREAK__":
            for col in (1, 2, 3):
                cell = ws.cell(row=row, column=col, value=None)
                cell.fill = rfill
                cell.border = CELL_BORDER
            ws.row_dimensions[row].height = 8
            row += 1
            continue

        style_cell(ws.cell(row=row, column=1, value=label),
                   fill=rfill, font=FONT_LABEL, align="left")
        style_cell(ws.cell(row=row, column=2, value=app_metrics[key]),
                   fill=rfill, font=FONT_VALUE_BOLD, num_format=fmt)
        style_cell(ws.cell(row=row, column=3, value=web_metrics[key]),
                   fill=rfill, font=FONT_VALUE_BOLD, num_format=fmt)

        if fmt:
            pct_rows_to_bar.append(row)
        row += 1

    for r in pct_rows_to_bar:
        bar_rule = DataBarRule(
            start_type="num", start_value=0,
            end_type="num",   end_value=110,
            color=COLOR_BAR_FILL,
            showValue=True,
        )
        ws.conditional_formatting.add(f"B{r}:C{r}", bar_rule)

    # Table 2 - QA Coverage
    row += 2
    section_header(ws, row, 1, 3, "  QA COVERAGE")
    row += 1

    style_cell(ws.cell(row=row, column=1, value="Metric"), fill=SUB_FILL, font=FONT_HEADER, align="left")
    style_cell(ws.cell(row=row, column=2, value="App"),    fill=SUB_FILL, font=FONT_HEADER)
    style_cell(ws.cell(row=row, column=3, value="Web"),    fill=SUB_FILL, font=FONT_HEADER)
    row += 1

    overview_keys = ["Total Modules", "Total Unique Events", "Total Parameters Checked"]
    for idx, key in enumerate(overview_keys):
        rfill = ROW_ODD_FILL if idx % 2 == 0 else ROW_EVEN_FILL
        style_cell(ws.cell(row=row, column=1, value=key),
                   fill=rfill, font=FONT_LABEL, align="left")
        style_cell(ws.cell(row=row, column=2, value=app_metrics[key]),
                   fill=rfill, font=FONT_VALUE_BOLD)
        style_cell(ws.cell(row=row, column=3, value=web_metrics[key]),
                   fill=rfill, font=FONT_VALUE_BOLD)
        row += 1

    # Table 3 - Per-module breakdown
    row += 2
    section_header(ws, row, 1, 13, "  PER-MODULE BREAKDOWN  (params)")
    row += 1

    ws.cell(row=row, column=1).fill = SUB_FILL
    ws.cell(row=row, column=1).border = CELL_BORDER

    style_cell(ws.cell(row=row, column=2, value="App"), fill=SUB_FILL, font=FONT_HEADER)
    ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=7)
    for c in range(3, 8):
        ws.cell(row=row, column=c).fill = SUB_FILL
        ws.cell(row=row, column=c).border = CELL_BORDER

    style_cell(ws.cell(row=row, column=8, value="Web"), fill=SUB_FILL, font=FONT_HEADER)
    ws.merge_cells(start_row=row, end_row=row, start_column=8, end_column=13)
    for c in range(9, 14):
        ws.cell(row=row, column=c).fill = SUB_FILL
        ws.cell(row=row, column=c).border = CELL_BORDER
    row += 1

    headers = ["Module",
               "Events", "Parameters", "PASS", "PARTIAL", "FAIL", "Pass %",
               "Events", "Parameters", "PASS", "PARTIAL", "FAIL", "Pass %"]
    for i, h in enumerate(headers, 1):
        style_cell(ws.cell(row=row, column=i, value=h), fill=SUB_FILL, font=FONT_HEADER)
    row += 1

    app_modules = set(app_mod_params["Module"]) if not app_mod_params.empty else set()
    web_modules = set(web_mod_params["Module"]) if not web_mod_params.empty else set()
    present     = app_modules | web_modules

    # Use input-file order, filtered down to modules that actually appear in
    # either platform. Anything unexpectedly missing from module_order goes
    # to the end (sorted, just so the failure mode is at least deterministic).
    all_modules = [m for m in module_order if m in present]
    leftover    = sorted(present - set(all_modules))
    all_modules.extend(leftover)

    app_lookup = app_mod_params.set_index("Module").to_dict(orient="index") if not app_mod_params.empty else {}
    web_lookup = web_mod_params.set_index("Module").to_dict(orient="index") if not web_mod_params.empty else {}

    pmp_start = row
    metric_keys = ["Events", "Parameters", "PASS", "PARTIAL", "FAIL", "Pass %"]
    for idx, mod in enumerate(all_modules):
        rfill = ROW_ODD_FILL if idx % 2 == 0 else ROW_EVEN_FILL
        a = app_lookup.get(mod, {})
        w = web_lookup.get(mod, {})

        style_cell(ws.cell(row=row, column=1, value=mod),
                   fill=rfill, font=FONT_VALUE_BOLD, align="left")

        for col, k in zip(range(2, 8), metric_keys):
            val = a.get(k, "" if k != "Pass %" else 0.0)
            fmt = '0.0"%"' if k == "Pass %" else None
            font = FONT_VALUE_BOLD if k == "Pass %" else FONT_VALUE
            if not a:
                val = "—"
                fmt = None
                font = FONT_VALUE
            style_cell(ws.cell(row=row, column=col, value=val),
                       fill=rfill, font=font, num_format=fmt)

        for col, k in zip(range(8, 14), metric_keys):
            val = w.get(k, "" if k != "Pass %" else 0.0)
            fmt = '0.0"%"' if k == "Pass %" else None
            font = FONT_VALUE_BOLD if k == "Pass %" else FONT_VALUE
            if not w:
                val = "—"
                fmt = None
                font = FONT_VALUE
            style_cell(ws.cell(row=row, column=col, value=val),
                       fill=rfill, font=font, num_format=fmt)

        row += 1
    pmp_end = row - 1

    if pmp_end >= pmp_start:
        bar_rule_app = DataBarRule(
            start_type="num", start_value=0,
            end_type="num",   end_value=110,
            color=COLOR_BAR_FILL,
            showValue=True,
        )
        ws.conditional_formatting.add(f"G{pmp_start}:G{pmp_end}", bar_rule_app)

        bar_rule_web = DataBarRule(
            start_type="num", start_value=0,
            end_type="num",   end_value=110,
            color=COLOR_BAR_FILL,
            showValue=True,
        )
        ws.conditional_formatting.add(f"M{pmp_start}:M{pmp_end}", bar_rule_web)

    last_used_row = row - 1

    widths = {"A": 26,
              "B": 9,  "C": 12, "D": 9,  "E": 10, "F": 8,  "G": 11,
              "H": 9,  "I": 12, "J": 9,  "K": 10, "L": 8,  "M": 11,
              "N": 4}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    visible_max_row = max(30, last_used_row + 2)
    for col_letter in ["O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]:
        ws.column_dimensions[col_letter].hidden = True
    for r in range(visible_max_row + 1, 101):
        ws.row_dimensions[r].hidden = True


# =========================
# CLI
# =========================
def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "GA4 event validation. Reads the rules + GA4 actual values from "
            "the same workbook and writes Validation + Summary sheets."
        ),
    )
    parser.add_argument(
        "--mode",
        choices=["normal", "ignore-required"],
        default="normal",
        help=(
            "normal (default): only required=yes rows are value-matched; "
            "non-required rows are marked Manual Check. "
            "ignore-required: every QA'd row is value-matched regardless of "
            "the required column — useful for quick coverage scans."
        ),
    )
    return parser.parse_args()


# =========================
# MAIN
# =========================
def main():
    args = parse_args()
    ignore_required = (args.mode == "ignore-required")
    mode_label      = "Ignore-Required" if ignore_required else "Normal"

    print(f"Mode: {mode_label}")
    print("Reading Rules file (with GA4 values)...\n")
    events, input_columns, module_order = read_excel(RULES_FILE, SHEET_NAME)

    print("\nLoaded combinations:")
    for v in events.values():
        print(f"  → [{v['module'] or 'Uncategorized'}] {v['event_name']} | "
              f"platform={v['platform']} | params={len(v['params'])}")

    print(f"\nModule order from input file ({len(module_order)} modules):")
    for i, m in enumerate(module_order, 1):
        print(f"  {i:>3}. {m}")

    result_df = validate(events, ignore_required=ignore_required)

    # Tag the output file with the mode so a normal run and an ignore-required
    # run don't overwrite each other.
    mode_tag = "_ignoreReq" if ignore_required else ""
    if KEEP_HISTORY:
        timestamp   = datetime.now(ZoneInfo('Asia/Kolkata')).strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_FILE.replace(".xlsx", f"{mode_tag}_{timestamp}.xlsx")
    else:
        output_path = OUTPUT_FILE.replace(".xlsx", f"{mode_tag}.xlsx") if mode_tag else OUTPUT_FILE

    app_metrics, app_mod_params = build_metrics_for_platform(
        result_df, "APP", TOTAL_EVENTS_APP, TOTAL_MODULES_APP, module_order)
    web_metrics, web_mod_params = build_metrics_for_platform(
        result_df, "WEB", TOTAL_EVENTS_WEB, TOTAL_MODULES_WEB, module_order)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ============================================================
        # Build the Validation sheet view:
        #   - Drop internal-only columns (prefixed with _ )
        #   - Place original input columns FIRST (preserving input order)
        #   - Append Status + Comments at the end
        # Row order is module → event → param as they appeared in the input
        # file (already enforced upstream by dict insertion + list append).
        # ============================================================
        internal_cols = [c for c in result_df.columns if c.startswith("_")]
        validation_view = result_df.drop(columns=internal_cols).copy()

        # Build the final column order: original input columns (those that exist
        # in the result), then any other non-internal cols (which will be Status
        # and Comments), preserving their order.
        original_present = [c for c in input_columns if c in validation_view.columns]
        extras = [c for c in validation_view.columns if c not in original_present]
        ordered_cols = original_present + extras
        validation_view = validation_view[ordered_cols]

        validation_view["Comments"] = validation_view["Comments"].fillna("").astype(str)
        validation_view.to_excel(writer, sheet_name="Validation", index=False)

        wb = writer.book
        val_ws = wb["Validation"]

        STATUS_COLOR_MAP = {
            "Pass":                 CF_GREEN_FILL,
            "Partial - Missing":    CF_YELLOW_FILL,
            "Partial - Additional": CF_YELLOW_FILL,
            "Failed":               CF_RED_FILL,
            # "Manual Check" intentionally omitted -> no fill
        }

        status_col_idx = list(validation_view.columns).index("Status") + 1
        for row_idx in range(2, len(validation_view) + 2):
            cell = val_ws.cell(row=row_idx, column=status_col_idx)
            fill = STATUS_COLOR_MAP.get(cell.value)
            if fill is not None:
                cell.fill = fill

        write_summary_sheet(wb, app_metrics, app_mod_params,
                            web_metrics, web_mod_params, module_order,
                            mode_label=mode_label)
        wb.move_sheet("Summary", offset=-wb.sheetnames.index("Summary"))

    print(f"\n{'='*60}")
    print(f"CONSOLE SUMMARY  (mode: {mode_label})")
    print(f"{'='*60}")

    # (not set) detection — single line, easy to spot
    if "_has_not_set" in result_df.columns:
        ns_total = int(result_df["_has_not_set"].sum())
        if ns_total:
            ns_by_plat = (result_df[result_df["_has_not_set"]]
                          .groupby("_platform").size().to_dict())
            breakdown = ", ".join(f"{p}: {n}" for p, n in sorted(ns_by_plat.items()))
            print(f"\n⚠ Rows containing (not set): {ns_total}  ({breakdown})")
        else:
            print("\n✓ No (not set) values detected in GA4 actual values.")
    for plat_label, m, mod_p in [("APP", app_metrics, app_mod_params),
                                  ("WEB", web_metrics, web_mod_params)]:
        print(f"\n{plat_label}")
        print(f"  Total Events                : {m['Total Events']}")
        print(f"  QA Completed (events)       : {m['Events Accounted']}")
        print(f"  QA Completion % (Events)    : {m['Event Completion %']}%")
        print(f"  Modules                     : {m['Modules']}")
        print(f"  QA Completed (modules)      : {m['Modules Worked On']}")
        print(f"  Completion % (Modules)      : {m['Module Progress %']}%")
        if not mod_p.empty:
            tp = int(mod_p["Parameters"].sum())
            tps = int(mod_p["PASS"].sum())
            assert tps + int(mod_p["PARTIAL"].sum()) + int(mod_p["FAIL"].sum()) == tp
    print(f"\nDone. Output saved to: {output_path}")


if __name__ == "__main__":
    main()