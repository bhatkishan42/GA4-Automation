# Version 19 — simplified status set: Pass / Failed / Partial - Missing / Partial - Additional / Manual Check

import pandas as pd
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule

# =========================
# CONFIG — only change these
# =========================
RULES_FILE  = "Test_file.xlsx"            # the Rules / spec workbook
SHEET_NAME  = "bqtest"
BQ_FILE     = "ga4_bq_events_merged.csv"  # the BQ data export
OUTPUT_FILE = "validation_output.xlsx"

# Set to "app", "web", or None to auto-detect from stream_name column in CSV
PLATFORM_OVERRIDE = None

# If True, appends timestamp to output filename (keeps run history)
KEEP_HISTORY = True

# Hardcoded total events & modules per platform — UPDATE WHEN FULL SPEC CHANGES
TOTAL_EVENTS_APP    = 275   # change when full app event count changes
TOTAL_EVENTS_WEB    = 178   # change when full web event count is known
TOTAL_MODULES_APP   = 29    # change when full app module count changes
TOTAL_MODULES_WEB   = 29    # change when full web module count changes

PLATFORM_MAP = {
    "web stream"       : "web",
    "core android app" : "app",
    "core ios app"     : "app",
}

# =========================
# Racing THEME COLORS (Summary sheet)
# =========================
COLOR_BG             = "15151E"   # Excel - Summary sheet background
COLOR_HEADER         = "FF1E00"   # Excel - section header bands & title bar (bright red)
COLOR_SUBHEADER      = "1E1E2A"   # Excel - sub-header rows & column headers
COLOR_ROW_ODD        = "1A1A27"   # Excel - odd data row fill
COLOR_ROW_EVEN       = "22222F"   # Excel - even data row fill
COLOR_TEXT_PRIMARY   = "FFFFFF"   # Excel - primary text (titles, values)
COLOR_TEXT_SECONDARY = "E8E8E8"   # Excel - secondary text (labels)
COLOR_TEXT_MUTED     = "C0C0C0"   # Excel - muted text (timestamp, subtitles)
COLOR_BORDER         = "3A3A4A"   # Excel - cell borders
COLOR_BAR_FILL       = "00B050"   # Excel - progress bar fill (green; renders as white→green gradient)

# Validation sheet — traffic-light colors painted directly on Status cells
COLOR_VAL_GREEN      = "63BE7B"   # Validation - Pass
COLOR_VAL_YELLOW     = "FFEB84"   # Validation - Partial - Missing / Partial - Additional
COLOR_VAL_RED        = "F8696B"   # Validation - Failed

BG_FILL       = PatternFill("solid", fgColor=COLOR_BG)
HEADER_FILL   = PatternFill("solid", fgColor=COLOR_HEADER)
SUB_FILL      = PatternFill("solid", fgColor=COLOR_SUBHEADER)
ROW_ODD_FILL  = PatternFill("solid", fgColor=COLOR_ROW_ODD)
ROW_EVEN_FILL = PatternFill("solid", fgColor=COLOR_ROW_EVEN)

CF_GREEN_FILL  = PatternFill("solid", fgColor=COLOR_VAL_GREEN)
CF_YELLOW_FILL = PatternFill("solid", fgColor=COLOR_VAL_YELLOW)
CF_RED_FILL    = PatternFill("solid", fgColor=COLOR_VAL_RED)

THIN_BORDER  = Side(border_style="thin", color=COLOR_BORDER)
CELL_BORDER  = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)

FONT_TITLE      = Font(name="Calibri", size=18, bold=True, color=COLOR_TEXT_PRIMARY)
FONT_SUBTITLE   = Font(name="Calibri", size=10, italic=True, color=COLOR_TEXT_MUTED)
FONT_SECTION    = Font(name="Calibri", size=12, bold=True, color=COLOR_TEXT_PRIMARY)
FONT_HEADER     = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT_PRIMARY)
FONT_LABEL      = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT_SECONDARY)
FONT_VALUE      = Font(name="Calibri", size=11, color=COLOR_TEXT_PRIMARY)
FONT_VALUE_BOLD = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT_PRIMARY)


# =========================
# NORMALIZE HELPER
# =========================
def normalize(value):
    return str(value).strip().lower()


# =========================
# READ EXCEL
# =========================
def read_excel(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    df.columns = df.columns.str.strip().str.lower()

    # AUTO-FIX swapped event_name/platform
    platform_labels = {"app", "web"}
    mask = df["event_name"].astype(str).str.strip().str.lower().isin(platform_labels)
    if mask.any():
        fixed_count = int(mask.sum())
        print(f"⚠ Auto-fixing {fixed_count} rows where event_name/platform are swapped")
        swapped_events   = df.loc[mask, "platform"].copy()
        swapped_platform = df.loc[mask, "event_name"].copy()
        df.loc[mask, "event_name"] = swapped_events
        df.loc[mask, "platform"]   = swapped_platform

    df = df[df["event_name"].notna() & (df["event_name"].astype(str).str.strip() != "")]

    events = {}
    duplicates = []

    for _, row in df.iterrows():
        raw_mod         = row.get("module", "")
        module          = "" if pd.isna(raw_mod) else str(raw_mod).strip()
        raw_event       = str(row["event_name"])
        platform        = normalize(row.get("platform", "all"))
        parameter_name  = normalize(row["parameter_name"])
        bq_column       = normalize(row["bq_column"])
        expected_raw    = "" if pd.isna(row["expected_value"]) else str(row["expected_value"])
        required        = str(row.get("required", "")).strip().lower() in ["yes", "true", "1"]
        expected_values = [normalize(v) for v in expected_raw.split("|") if v.strip()]

        event_names = [normalize(e) for e in raw_event.split("|") if e.strip()]

        for event_name in event_names:
            key = f"{event_name}||{platform}"

            if key not in events:
                events[key] = {
                    "module"     : module,
                    "event_name" : event_name,
                    "platform"   : platform,
                    "columns"    : [],
                    "params"     : []
                }

            existing_params = {p["parameter_name"] for p in events[key]["params"]}
            if parameter_name in existing_params:
                duplicates.append({
                    "event"     : event_name,
                    "platform"  : platform,
                    "parameter" : parameter_name,
                })
                continue

            if bq_column not in events[key]["columns"]:
                events[key]["columns"].append(bq_column)

            events[key]["params"].append({
                "parameter_name" : parameter_name,
                "bq_column"      : bq_column,
                "expected_values": expected_values,
                "required"       : required,
            })

    if duplicates:
        print(f"\n⚠ Found {len(duplicates)} duplicate parameter definitions (kept first occurrence):")
        for d in duplicates:
            print(f"    - {d['event']} [{d['platform']}] → {d['parameter']}")
        print()

    return events


# =========================
# READ CSV
# =========================
def read_csv(file_path):
    df = pd.read_csv(file_path)
    df.columns = df.columns.str.strip().str.lower()

    if PLATFORM_OVERRIDE:
        df["platform_group"] = PLATFORM_OVERRIDE
        print(f"Platform override: {PLATFORM_OVERRIDE}")
    elif "stream_name" in df.columns:
        df["platform_group"] = (
            df["stream_name"]
            .apply(normalize)
            .map(PLATFORM_MAP)
            .fillna("unknown")
        )
    else:
        df["platform_group"] = "unknown"

    print(f"BQ loaded       : {len(df)} rows")
    print(f"BQ columns      : {list(df.columns)}")
    print(f"Platforms found : {df['platform_group'].unique().tolist()}\n")
    return df


# =========================
# VALIDATE
# =========================
def validate(events, df):
    results = []

    for key, config in events.items():
        module     = config["module"]
        event_name = config["event_name"]
        platform   = config["platform"]

        event_df = df[df["event_name"].apply(normalize) == event_name]

        if platform != "all":
            event_df = event_df[event_df["platform_group"] == platform]

        event_found = len(event_df) > 0

        print(f"\n▶ [{module or 'Uncategorized'}] {event_name} [{platform.upper()}]"
              f" — {'FOUND' if event_found else 'NOT FOUND'} ({len(event_df)} rows)")

        for param in config["params"]:
            param_name      = param["parameter_name"]
            bq_col          = param["bq_column"]
            expected_values = param["expected_values"]
            required        = param["required"]
            required_label  = "YES" if required else "NO"

            if not required:
                status = "NOT REQUIRED"
                actual_values_sorted = []
                print(f"  - {param_name} | NOT REQUIRED — skipped")

            elif not event_found:
                status = "EVENT NOT FOUND"
                actual_values_sorted = []

            elif bq_col not in event_df.columns:
                status = "COLUMN NOT IN CSV"
                actual_values_sorted = []

            else:
                actual_values_set = set(
                    event_df[bq_col]
                    .dropna()
                    .apply(normalize)
                    .unique()
                )
                actual_values_set.discard("")
                actual_values_sorted = sorted(actual_values_set)

                if not actual_values_set:
                    status = "PARAMETER MISSING"
                elif not expected_values:
                    status = "FOUND (no expected value specified)"
                elif all(v in actual_values_set for v in expected_values):
                    additional = sorted([v for v in actual_values_sorted if v not in expected_values])
                    status = f"PARTIAL — additional values: {', '.join(additional)}" if additional else "PASS"
                elif any(v in actual_values_set for v in expected_values):
                    missing = [v for v in expected_values if v not in actual_values_set]
                    status  = f"PARTIAL — missing: {', '.join(missing)}"
                else:
                    status = "FAIL — no expected values found in CSV"

            if required:
                icon = (
                    "✔" if status == "PASS"
                    else "~" if any(x in status for x in ["no expected", "PARTIAL"])
                    else "✖"
                )
                print(f"  {icon} {param_name} | {status}")

            display_status, comments = status_to_columns(status)

            results.append({
                "module"              : module,
                "event_name"          : event_name,
                "platform"            : platform.upper(),
                "event_found_in_csv"  : "YES" if event_found else "NO",
                "parameter_name"      : param_name,
                "bq_column"           : bq_col,
                "required"            : required_label,
                "expected_value"      : "|".join(expected_values) if expected_values else "(any)",
                "actual_values_in_csv": "|".join(actual_values_sorted),
                "Status"              : display_status,
                "Comments"            : comments,
                "status"              : status,   # internal — kept for categorize() & summary logic
            })

    return pd.DataFrame(results)


# =========================
# STATUS / COMMENTS TRANSLATOR
# =========================
def status_to_columns(internal_status):
    """
    Translate the internal status string used by validate() / categorize() into
    a (display_status, comments) pair for the Validation output sheet.

    Display values (only 5):
      - Pass
      - Failed
      - Partial - Missing
      - Partial - Additional
      - Manual Check

    "Found" / "Event Missing" / "Parameter Missing" cases all collapse to
    "Failed" with a descriptive comment so the user can still tell them apart.
    """
    s = internal_status

    # Pass
    if s == "PASS":
        return "Pass", ""

    # Manual Check (was Not Required — skipped, awaiting human review)
    if s == "NOT REQUIRED":
        return "Manual Check", ""

    # Failed buckets — all show "Failed" with the reason in Comments
    if s == "EVENT NOT FOUND":
        return "Failed", "event not in BQ"
    if s == "PARAMETER MISSING":
        return "Failed", "no values in BQ"
    if s == "COLUMN NOT IN CSV":
        return "Failed", "column not in BQ"
    if s == "FAIL — no expected values found in CSV":
        return "Failed", "no expected values found in BQ"
    if s == "FOUND (no expected value specified)":
        return "Failed", "data exists but no expected value specified in Rules"

    # Partial variants
    if s.startswith("PARTIAL — missing: "):
        return "Partial - Missing", s.replace("PARTIAL — ", "")
    if s.startswith("PARTIAL — additional values: "):
        return "Partial - Additional", s.replace("PARTIAL — ", "")

    # Fallback — anything unexpected lands in Failed with the raw reason
    return "Failed", s


# =========================
# SUMMARY HELPERS
# =========================
def categorize(status):
    """Bucket the internal status for summary calculations."""
    if status == "PASS" or status == "FOUND (no expected value specified)":
        # NOTE: "Found" is shown as Failed in the Validation sheet, but for
        # summary calculations we still treat it as PASS-adjacent (data
        # exists). If you'd rather summary count it as FAIL too, change
        # the line above to: if status == "PASS":
        return "PASS"
    if status.startswith("PARTIAL"):
        return "PARTIAL"
    if status == "NOT REQUIRED":
        return "EXCLUDED"
    return "FAIL"


def build_metrics_for_platform(result_df, platform_label, total_events_spec, total_modules_spec):
    """Compute completion-rate metrics + per-module param table for one platform."""
    df = result_df[result_df["platform"].str.upper() == platform_label].copy()
    df["module_label"] = df["module"].replace("", "Uncategorized").fillna("Uncategorized")
    df["bucket"]       = df["status"].apply(categorize)

    validated   = df[df["bucket"] != "EXCLUDED"]
    events_acc  = df["event_name"].nunique() if len(df) else 0
    modules_acc = df["module_label"].nunique() if len(df) else 0

    modules_worked_on = validated["module_label"].nunique() if len(validated) else 0

    required = df[df["required"] == "YES"]
    modules_completed = 0
    if not required.empty:
        for _, grp in required.groupby("module_label"):
            if (grp["bucket"] == "PASS").all():
                modules_completed += 1

    metrics = {
        "Total Events"         : total_events_spec,
        "Events Accounted"     : events_acc,
        "Event Completion %"   : round(events_acc / total_events_spec * 100, 2) if total_events_spec else 0.0,
        "Modules"              : total_modules_spec,
        "Modules Worked On"    : modules_worked_on,
        "Module Progress %"    : round(modules_worked_on / total_modules_spec * 100, 2) if total_modules_spec else 0.0,
        "Modules Completed"    : modules_completed,
        "Module Completion %"  : round(modules_completed / total_modules_spec * 100, 2) if total_modules_spec else 0.0,
        "Total Modules"        : modules_acc,
        "Total Unique Events"  : events_acc,
        "Total Parameters Checked" : len(validated),
    }

    rows = []
    for mod, grp in validated.groupby("module_label"):
        n  = len(grp)
        p  = int((grp["bucket"] == "PASS").sum())
        pp = int((grp["bucket"] == "PARTIAL").sum())
        fl = int((grp["bucket"] == "FAIL").sum())
        events_in_mod = grp["event_name"].nunique()
        assert p + pp + fl == n, f"Math mismatch for {mod}: {p}+{pp}+{fl}!={n}"
        rows.append({
            "Module"     : mod,
            "Events"     : events_in_mod,
            "Parameters" : n,
            "PASS"       : p,
            "PARTIAL"    : pp,
            "FAIL"       : fl,
            "Pass %"     : round(p / n * 100, 1) if n else 0.0,
        })
    module_param_df = pd.DataFrame(rows).sort_values("Module").reset_index(drop=True) if rows else pd.DataFrame()

    return metrics, module_param_df


# =========================
# WRITE SUMMARY SHEET
# =========================
def fill_background(ws, max_row, max_col):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = BG_FILL


def section_header(ws, row, col_start, col_end, text):
    """Render a red section header band across given column range."""
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = FONT_SECTION
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if col_end > col_start:
        ws.merge_cells(start_row=row, end_row=row, start_column=col_start, end_column=col_end)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = HEADER_FILL
    ws.row_dimensions[row].height = 24


def style_cell(cell, value=None, fill=None, font=None, border=CELL_BORDER, align="center", num_format=None):
    if value is not None:
        cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if border is not None:
        cell.border = border
    if align == "center":
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif align == "left":
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if num_format:
        cell.number_format = num_format


def write_summary_sheet(wb, app_metrics, app_mod_params, web_metrics, web_mod_params):
    """Single Summary sheet — three tables, App | Web side-by-side layout."""
    sheet_name = "Summary"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    fill_background(ws, max_row=80, max_col=14)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COLOR_HEADER

    # ---- TITLE ----
    style_cell(ws.cell(row=1, column=1, value="BQ EVENT VALIDATION  |  SUMMARY"),
               fill=HEADER_FILL, font=FONT_TITLE, border=None)
    ws.merge_cells("A1:N1")
    ws.row_dimensions[1].height = 40

    style_cell(ws.cell(row=2, column=1,
               value=f"Generated: {datetime.now(ZoneInfo('Asia/Kolkata')).strftime('%B %d, %Y at %I:%M %p IST')}"),
               fill=BG_FILL, font=FONT_SUBTITLE, border=None)
    ws.merge_cells("A2:N2")
    ws.row_dimensions[2].height = 20

    # ============================================================
    # Table 1 - QA Overview (App | Web side by side)
    # ============================================================
    row = 4
    section_header(ws, row, 1, 3, "  QA OVERVIEW")
    row += 1

    style_cell(ws.cell(row=row, column=1, value=""),    fill=SUB_FILL, font=FONT_HEADER, align="left")
    style_cell(ws.cell(row=row, column=2, value="App"), fill=SUB_FILL, font=FONT_HEADER)
    style_cell(ws.cell(row=row, column=3, value="Web"), fill=SUB_FILL, font=FONT_HEADER)
    row += 1

    pct_rows_to_bar = []
    cr_metrics = [
        ("Total Events",              "Total Events",        None),
        ("QA Completed",              "Events Accounted",    None),
        ("QA Completion % (Events)",  "Event Completion %",  '0.00"%"'),
        ("__BREAK__",                 None,                  None),
        ("Modules",                   "Modules",             None),
        ("QA Completed",              "Modules Worked On",   None),
        ("Completion % (Modules)",    "Module Progress %",   '0.00"%"'),
    ]

    for idx, (label, key, fmt) in enumerate(cr_metrics):
        rfill = ROW_ODD_FILL if idx % 2 == 0 else ROW_EVEN_FILL

        if label == "__BREAK__":
            for col in (1, 2, 3):
                cell = ws.cell(row=row, column=col, value=None)
                cell.fill = rfill
                cell.border = CELL_BORDER
            ws.row_dimensions[row].height = 8
            row += 1
            continue

        style_cell(ws.cell(row=row, column=1, value=label),
                   fill=rfill, font=FONT_LABEL, align="left")
        style_cell(ws.cell(row=row, column=2, value=app_metrics[key]),
                   fill=rfill, font=FONT_VALUE_BOLD, num_format=fmt)
        style_cell(ws.cell(row=row, column=3, value=web_metrics[key]),
                   fill=rfill, font=FONT_VALUE_BOLD, num_format=fmt)

        if fmt:
            pct_rows_to_bar.append(row)
        row += 1

    for r in pct_rows_to_bar:
        bar_rule = DataBarRule(
            start_type="num", start_value=0,
            end_type="num",   end_value=110,
            color=COLOR_BAR_FILL,
            showValue=True,
        )
        ws.conditional_formatting.add(f"B{r}:C{r}", bar_rule)

    # ============================================================
    # Table 2 - QA Coverage
    # ============================================================
    row += 2
    section_header(ws, row, 1, 3, "  QA COVERAGE")
    row += 1

    style_cell(ws.cell(row=row, column=1, value="Metric"), fill=SUB_FILL, font=FONT_HEADER, align="left")
    style_cell(ws.cell(row=row, column=2, value="App"),    fill=SUB_FILL, font=FONT_HEADER)
    style_cell(ws.cell(row=row, column=3, value="Web"),    fill=SUB_FILL, font=FONT_HEADER)
    row += 1

    overview_keys = ["Total Modules", "Total Unique Events", "Total Parameters Checked"]
    for idx, key in enumerate(overview_keys):
        rfill = ROW_ODD_FILL if idx % 2 == 0 else ROW_EVEN_FILL
        style_cell(ws.cell(row=row, column=1, value=key),
                   fill=rfill, font=FONT_LABEL, align="left")
        style_cell(ws.cell(row=row, column=2, value=app_metrics[key]),
                   fill=rfill, font=FONT_VALUE_BOLD)
        style_cell(ws.cell(row=row, column=3, value=web_metrics[key]),
                   fill=rfill, font=FONT_VALUE_BOLD)
        row += 1

    # ============================================================
    # Table 3 - Per-module breakdown (params), App | Web side by side
    # ============================================================
    row += 2
    section_header(ws, row, 1, 13, "  PER-MODULE BREAKDOWN  (params)")
    row += 1

    ws.cell(row=row, column=1).fill = SUB_FILL
    ws.cell(row=row, column=1).border = CELL_BORDER

    style_cell(ws.cell(row=row, column=2, value="App"), fill=SUB_FILL, font=FONT_HEADER)
    ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=7)
    for c in range(3, 8):
        ws.cell(row=row, column=c).fill = SUB_FILL
        ws.cell(row=row, column=c).border = CELL_BORDER

    style_cell(ws.cell(row=row, column=8, value="Web"), fill=SUB_FILL, font=FONT_HEADER)
    ws.merge_cells(start_row=row, end_row=row, start_column=8, end_column=13)
    for c in range(9, 14):
        ws.cell(row=row, column=c).fill = SUB_FILL
        ws.cell(row=row, column=c).border = CELL_BORDER
    row += 1

    headers = ["Module",
               "Events", "Parameters", "PASS", "PARTIAL", "FAIL", "Pass %",
               "Events", "Parameters", "PASS", "PARTIAL", "FAIL", "Pass %"]
    for i, h in enumerate(headers, 1):
        style_cell(ws.cell(row=row, column=i, value=h), fill=SUB_FILL, font=FONT_HEADER)
    row += 1

    app_modules = set(app_mod_params["Module"]) if not app_mod_params.empty else set()
    web_modules = set(web_mod_params["Module"]) if not web_mod_params.empty else set()
    all_modules = sorted(app_modules | web_modules)

    app_lookup = app_mod_params.set_index("Module").to_dict(orient="index") if not app_mod_params.empty else {}
    web_lookup = web_mod_params.set_index("Module").to_dict(orient="index") if not web_mod_params.empty else {}

    pmp_start = row
    metric_keys = ["Events", "Parameters", "PASS", "PARTIAL", "FAIL", "Pass %"]
    for idx, mod in enumerate(all_modules):
        rfill = ROW_ODD_FILL if idx % 2 == 0 else ROW_EVEN_FILL
        a = app_lookup.get(mod, {})
        w = web_lookup.get(mod, {})

        style_cell(ws.cell(row=row, column=1, value=mod),
                   fill=rfill, font=FONT_VALUE_BOLD, align="left")

        for col, k in zip(range(2, 8), metric_keys):
            val = a.get(k, "" if k != "Pass %" else 0.0)
            fmt = '0.0"%"' if k == "Pass %" else None
            font = FONT_VALUE_BOLD if k == "Pass %" else FONT_VALUE
            if not a:
                val = "—"
                fmt = None
                font = FONT_VALUE
            style_cell(ws.cell(row=row, column=col, value=val),
                       fill=rfill, font=font, num_format=fmt)

        for col, k in zip(range(8, 14), metric_keys):
            val = w.get(k, "" if k != "Pass %" else 0.0)
            fmt = '0.0"%"' if k == "Pass %" else None
            font = FONT_VALUE_BOLD if k == "Pass %" else FONT_VALUE
            if not w:
                val = "—"
                fmt = None
                font = FONT_VALUE
            style_cell(ws.cell(row=row, column=col, value=val),
                       fill=rfill, font=font, num_format=fmt)

        row += 1
    pmp_end = row - 1

    if pmp_end >= pmp_start:
        bar_rule_app = DataBarRule(
            start_type="num", start_value=0,
            end_type="num",   end_value=110,
            color=COLOR_BAR_FILL,
            showValue=True,
        )
        ws.conditional_formatting.add(f"G{pmp_start}:G{pmp_end}", bar_rule_app)

        bar_rule_web = DataBarRule(
            start_type="num", start_value=0,
            end_type="num",   end_value=110,
            color=COLOR_BAR_FILL,
            showValue=True,
        )
        ws.conditional_formatting.add(f"M{pmp_start}:M{pmp_end}", bar_rule_web)

    last_used_row = row - 1

    widths = {"A": 26,
              "B": 9,  "C": 12, "D": 9,  "E": 10, "F": 8,  "G": 11,
              "H": 9,  "I": 12, "J": 9,  "K": 10, "L": 8,  "M": 11,
              "N": 4}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    visible_max_row = max(30, last_used_row + 2)
    for col_letter in ["O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]:
        ws.column_dimensions[col_letter].hidden = True
    for r in range(visible_max_row + 1, 101):
        ws.row_dimensions[r].hidden = True


# =========================
# MAIN
# =========================
def main():
    print("Reading Rules file...\n")
    events = read_excel(RULES_FILE, SHEET_NAME)

    print("\nLoaded combinations:")
    for v in events.values():
        print(f"  → [{v['module'] or 'Uncategorized'}] {v['event_name']} | "
              f"platform={v['platform']} | params={len(v['params'])}")

    print("\nReading BQ file...\n")
    df = read_csv(BQ_FILE)

    result_df = validate(events, df)

    if KEEP_HISTORY:
        timestamp   = datetime.now(ZoneInfo('Asia/Kolkata')).strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_FILE.replace(".xlsx", f"_{timestamp}.xlsx")
    else:
        output_path = OUTPUT_FILE

    app_metrics, app_mod_params = build_metrics_for_platform(
        result_df, "APP", TOTAL_EVENTS_APP, TOTAL_MODULES_APP)
    web_metrics, web_mod_params = build_metrics_for_platform(
        result_df, "WEB", TOTAL_EVENTS_WEB, TOTAL_MODULES_WEB)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        validation_view = result_df.drop(columns=["status"]).copy()
        validation_view["Comments"] = validation_view["Comments"].fillna("").astype(str)
        validation_view.to_excel(writer, sheet_name="Validation", index=False)

        wb = writer.book
        val_ws = wb["Validation"]

        # ============================================================
        # Direct fill the Status column based on value (no conditional
        # formatting — openpyxl's CF support is unreliable across Excel
        # versions, so we just paint the cells ourselves). Reliable, but
        # static: if you hand-edit a Status cell in Excel later, the colour
        # won't auto-update. Re-running the script regenerates everything.
        # ============================================================
        STATUS_COLOR_MAP = {
            "Pass":                 CF_GREEN_FILL,
            "Partial - Missing":    CF_YELLOW_FILL,
            "Partial - Additional": CF_YELLOW_FILL,
            "Failed":               CF_RED_FILL,
            # "Manual Check" intentionally omitted -> no fill
        }

        status_col_idx = list(validation_view.columns).index("Status") + 1
        for row_idx in range(2, len(validation_view) + 2):  # row 1 is header
            cell = val_ws.cell(row=row_idx, column=status_col_idx)
            fill = STATUS_COLOR_MAP.get(cell.value)
            if fill is not None:
                cell.fill = fill

        write_summary_sheet(wb, app_metrics, app_mod_params, web_metrics, web_mod_params)
        wb.move_sheet("Summary", offset=-wb.sheetnames.index("Summary"))

    print(f"\n{'='*60}")
    print("CONSOLE SUMMARY")
    print(f"{'='*60}")
    for plat_label, m, mod_p in [("APP", app_metrics, app_mod_params),
                                  ("WEB", web_metrics, web_mod_params)]:
        print(f"\n{plat_label}")
        print(f"  Total Events                : {m['Total Events']}")
        print(f"  QA Completed (events)       : {m['Events Accounted']}")
        print(f"  QA Completion % (Events)    : {m['Event Completion %']}%")
        print(f"  Modules                     : {m['Modules']}")
        print(f"  QA Completed (modules)      : {m['Modules Worked On']}")
        print(f"  Completion % (Modules)      : {m['Module Progress %']}%")
        if not mod_p.empty:
            tp = int(mod_p["Parameters"].sum())
            tps = int(mod_p["PASS"].sum())
            assert tps + int(mod_p["PARTIAL"].sum()) + int(mod_p["FAIL"].sum()) == tp
    print(f"\nDone. Output saved to: {output_path}")


if __name__ == "__main__":
    main()